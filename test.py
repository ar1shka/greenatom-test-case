import os
import smtplib
import pandas as pd
import numpy as np
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

SERVER_ADDRESS = os.environ.get("SERVER_ADDRESS", "smtp.mail.ru")
SERVER_PORT = os.environ.get("SERVER_PORT", 465)

LOGIN = os.environ.get("LOGIN", None)
PASSWORD = os.environ.get("PASSWORD", None)

# Парсинг XML-файла
def parse_xml(file_path: str) :
    tree = ET.parse(file_path)
    root = tree.getroot()
    
    rows = list()

    for row in root[0][1].findall("row"):
        attributes = row.attrib
        if attributes['clearing'] == 'vk':
            secid = attributes['secid']
            rows.append([
                attributes[column]
                for column in ['tradedate', 'rate', 'tradetime']
            ])

    columns = [
        f"{column} {secid}"
        for column in ['Дата', 'Курс', 'Время']
    ]
    
    df = pd.DataFrame(
        rows,
        columns=columns
    )

    df[f"Курс {secid}"] = df[f"Курс {secid}"].astype(np.double)

    return df

# Проверка склонения слова "строка"
def checkstr(file_path: str):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    rows_m = sheet.max_row
    reply = f"В этом документе {rows_m} строк"
    rows_ost = rows_m % 10
    

    if rows_ost == 0:
        return reply
    elif rows_m > 4 and rows_m < 20:
        return reply
    elif rows_ost == 1:
        return reply + "а"
    else:
        return reply + "и"

# Отправка сообщения
def send_message_to_me(file_path: str):
    server_address = SERVER_ADDRESS
    server_port = SERVER_PORT
    login, password = LOGIN, PASSWORD 

    msg = MIMEMultipart()
    msg['From'], msg['To'], msg['Subject'] = login, login, "Тестовое задание"
    msg.attach(MIMEText(checkstr(file_path), 'plain'))
    
    with open(file_path, "rb") as file:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(file.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f"attachment; filename={file_path}")

    msg.attach(part)
    with smtplib.SMTP_SSL(server_address, server_port) as server:
        server.login(login, password)
        server.send_message(msg, login, login)

# Настройка ширины ячейки относительно заполненности ячейки
def auto_weith(file_path: str):
    wb = load_workbook(file_path)
    sheet = wb.active

    for col in sheet.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[col[0].column_letter].width = adjusted_width
    wb.save(file_path)

# Установка финансового формата ячеек
def finance_format(file_path: str):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Установка финансового формата на ячейки со стоимостью.
    for column in ["B", "E", "G"]:
        for cell in sheet[column]:
            cell.number_format = '_("$"* #,##0.00_)_("$"* \(#,##0.00\)_("$"* "-"??_)_(@_)' # Является одним из финансовых форматов
    workbook.save(file_path)


if __name__ == "__main__":
    file_path = "task.xlsx"
    
    df_usd = parse_xml('./data/currencyRate-USD_RUB-20241216-20250116.xml')
    df_jup = parse_xml('./data/currencyRate-JPY_RUB-20241216-20250116.xml')

    df = pd.concat((df_usd, df_jup), axis=1)
    df['Результат'] = df['Курс USD/RUB'] / df['Курс JPY/RUB']
    df.to_excel(file_path, index=False)

    finance_format(file_path)
    auto_weith(file_path)
    
    if LOGIN and PASSWORD:
        send_message_to_me(file_path)
        